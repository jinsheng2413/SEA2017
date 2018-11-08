# -*- coding: utf-8 -*-

"""
@author: 郭春彪
@license: (C) Copyright 2018, Nari.
@file: xls_rw.py
@time: 2018/5/19 0019 11:52
@desc:
1.方法名：write
         说明：按单元格写入
        :param i: excel具体列数
        :param l: excel具体行数
        :param value: 要写入的内容
2.方法名：read
        说明：按单元格读取excel
        :param i: excel具体列数
        :param l: excel具体行数
        :return: 返回单元格读取内容
3.方法名：read_row_out
        说明：按行读取excel每一行数据
        :param row:
        :return: 返回读取一行的列表
4.方法名：write_into_row
        说明按行写入excel表格
        :param row: 具体哪一行
        :param list: 要写入的一行数据列表
5.方法名：close
          说明：关闭excel

"""
from openpyxl import load_workbook


class XlsRw(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        # 获取sheet页
        self.ws = self.wb.worksheets[0]
        # 获取最大行数
        self.r_row = self.wb.worksheets[0].max_row
        # 获取最大列数
        self.c_col = self.wb.worksheets[0].max_column

    def write(self, row, col, value):
        """
         方法名：write
         说明：按单元格写入
        :param row: excel具体列数
        :param col: excel具体行数
        :param value: 要写入的内容
        """
        # eg: coord:A1
        self.ws.cell(row, col).value = value
        self.wb.save(self.filename)

    def read(self, i, l):
        """
        方法名：read
        说明：按单元格读取excel
        :param i: excel具体列数
        :param l: excel具体行数
        :return: 返回单元格读取内容
        """
        val = self.ws.cell(i, l).value
        return val

    def read_row_out(self, row):
        """
        方法名：read_row_out
        说明：按行读取excel每一行数据
        :param row:
        :return: 返回读取一行的列表
        """
        line = self.ws.max_column

        ls = []
        for i in range(1, line + 1):
            val = self.ws.cell(row, i).value
            ls.append(val)

        return ls

    # 按行写入表格
    def write_into_row(self, row, ls):
        """
        方法名：write_into_row
        说明按行写入excel表格
        :param row: 具体哪一行
        :param ls: 要写入的一行数据列表
        """
        self.ws.cell(row, 1).value = ls[0]
        for i in range(1, len(ls)):
            self.ws.cell(row, i + 1).value = ls[i]

        self.wb.save(self.filename)
        # 获取最大行数

    def excel_table(self, filepath):
        """
        excel参数化时应用
        :param filepath:excl地址
        :return:返回读取数据：[{},{}]
        """
        t = XlsRw(filepath)
        lister = []

        for rownumber in range(1, t.r_row + 1):
            app = {}
            for i in range((len(t.read_row_out(rownumber)))):
                app.setdefault(t.read_row_out(
                    1)[i], t.read_row_out(rownumber)[i])

            lister.append(app)
        del lister[0]

        return lister

    def close(self):
        """
        方法名：close
          说明：关闭excel
        """
        self.wb.close()


if __name__ == '__main__':
    p = XlsRw('D:/excel/text.xlsx')
    print(p.write(4, 5, '测试电视剧覅'))
